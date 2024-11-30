from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from collections import defaultdict
from io import BytesIO


def index(request):
    template_name = 'index.html'

    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Nenhum arquivo selecionado')
            return redirect('index')

        if not file.name.endswith('.xlsx'):
            messages.error(request, 'Por favor, envie apenas arquivos .xlsx')
            return redirect('index')

        try:
            # Processar o arquivo
            output = processar_planilha(file)

            # Retornar o arquivo processado
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=processado_{file.name}'
            return response

        except Exception as e:
            messages.error(request, f'Erro ao processar arquivo: {str(e)}')
            return redirect('index')

    return render(request, template_name)


def processar_planilha(arquivo_entrada):
    # Carregar a planilha da memória
    wb = load_workbook(filename=arquivo_entrada)
    ws = wb.active

    # Dicionário para armazenar os dados agrupados
    dados_por_categoria = defaultdict(list)

    # Ler os dados e agrupar por categoria
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 4 or not all(row[:4]):
            continue
        categoria, produto, quantidade, preco = row[:4]
        try:
            quantidade = float(quantidade)
            preco = float(preco)
            dados_por_categoria[categoria].append({
                'produto': produto,
                'quantidade': quantidade,
                'preco': preco,
                'total': quantidade * preco
            })
        except (ValueError, TypeError):
            continue

    # Criar nova planilha
    new_wb = Workbook()
    new_ws = new_wb.active

    # Definir estilos
    header_font = Font(bold=True)
    categoria_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    subtotal_font = Font(bold=True)

    # Adicionar cabeçalhos
    headers = ['Categoria', 'Produto', 'Quantidade', 'Preço', 'Total']
    for col, header in enumerate(headers, 1):
        cell = new_ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font

    current_row = 2

    # Processar cada categoria
    for categoria in sorted(dados_por_categoria.keys()):
        produtos = dados_por_categoria[categoria]

        # Linha da categoria
        categoria_row = current_row
        for col in range(1, 6):
            cell = new_ws.cell(row=categoria_row, column=col)
            cell.fill = categoria_fill

        new_ws.cell(row=categoria_row, column=1, value=categoria)
        current_row += 1

        # Adicionar produtos
        subtotal = 0
        for produto in produtos:
            row_data = [
                categoria,
                produto['produto'],
                produto['quantidade'],
                produto['preco'],
                produto['total']
            ]
            for col, value in enumerate(row_data, 1):
                new_ws.cell(row=current_row, column=col, value=value)
            subtotal += produto['total']
            current_row += 1

        # Adicionar subtotal
        new_ws.cell(row=current_row, column=1, value='Subtotal')
        subtotal_cell = new_ws.cell(row=current_row, column=5, value=subtotal)
        subtotal_cell.font = subtotal_font

        current_row += 2

    # Ajustar largura das colunas
    for col in range(1, 6):
        max_length = 0
        column = new_ws.column_dimensions[chr(64 + col)]

        for cell in new_ws[chr(64 + col)]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        column.width = max_length + 2

    # Salvar em memória
    output = BytesIO()
    new_wb.save(output)
    output.seek(0)
    return output
